"""
Creator CRUD & CSV import — data management endpoints.

Endpoints
---------
  POST   /kols/add                   Add a single KOL
  POST   /kols/import-csv            Bulk import from CSV
  PUT    /kols/{id}                  Update a KOL (auto-snapshots old values)
  DELETE /kols/{id}                  Delete a KOL
  GET    /kols/template              Download CSV template
  GET    /kols/export                Export all KOLs as CSV
  GET    /kols/{id}/history          Metric history snapshots for a KOL
  POST   /kols/{id}/simulate-update  Simulate a TikTok API refresh (random drift)

All writes persist to data/sample_kols.json immediately.
KOL metric history persists to data/kol_history.json.
"""
import csv
import io
import json
import os
import random
import re
from datetime import datetime
from typing import List, Optional

from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Depends
from pydantic import BaseModel, field_validator

from backend import tenancy
from backend.auth import require_tenant

router = APIRouter()

# Data file locations are resolved per-tenant at call time (see backend/tenancy.py),
# so every endpoint reads/writes only the current merchant's isolated folder.

# ── KOL history helpers ───────────────────────────────────────
_SNAPSHOT_FIELDS = ("engagement_rate", "fit_score", "followers", "commission_rate", "avg_views", "avg_likes")


def _load_history() -> List[dict]:
    path = tenancy.history_path()
    if not os.path.exists(path):
        return []
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def _save_history(data: List[dict]):
    with open(tenancy.history_path(), "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def _record_snapshot(kol_record: dict):
    """Append a timestamped snapshot of the KOL's current metrics to history."""
    history = _load_history()
    entry = next((h for h in history if h["kol_id"] == kol_record["id"]), None)
    if entry is None:
        entry = {"kol_id": kol_record["id"], "snapshots": []}
        history.append(entry)

    snapshot = {"recorded_at": datetime.utcnow().isoformat()}
    for f in _SNAPSHOT_FIELDS:
        if f in kol_record and kol_record[f] is not None:
            snapshot[f] = kol_record[f]
    entry["snapshots"].append(snapshot)
    _save_history(history)

VALID_COUNTRIES = ["MY", "ID", "TH", "PH", "SG", "VN"]
VALID_CATEGORIES = ["beauty", "fashion", "home", "fmcg"]

CSV_TEMPLATE_HEADER = [
    "name", "country", "category", "followers",
    "engagement_rate", "fit_score", "commission_rate", "cost",
    "avg_views", "avg_likes", "gender_ratio", "age_group",
    "tiktok_url",
]


# ════════════════════════════════════════════════════════════════
#  Pydantic models
# ════════════════════════════════════════════════════════════════
class KOLCreate(BaseModel):
    """Fields required to add a new KOL."""
    name: str
    country: str = "MY"
    category: str = "beauty"
    followers: int
    engagement_rate: float
    fit_score: float = 0.75
    commission_rate: float = 0.15
    cost: float = 0.0
    avg_views: Optional[int] = None
    avg_likes: Optional[int] = None
    gender_ratio: Optional[float] = None
    age_group: Optional[str] = None
    tiktok_url: Optional[str] = None

    @field_validator("country")
    @classmethod
    def validate_country(cls, v):
        if v not in VALID_COUNTRIES:
            raise ValueError(f"country must be one of: {', '.join(VALID_COUNTRIES)}")
        return v

    @field_validator("category")
    @classmethod
    def validate_category(cls, v):
        v = v.lower()
        if v not in VALID_CATEGORIES:
            raise ValueError(f"category must be one of: {', '.join(VALID_CATEGORIES)}")
        return v

    @field_validator("followers")
    @classmethod
    def validate_followers(cls, v):
        if v < 0:
            raise ValueError("followers must be non-negative")
        return v

    @field_validator("cost")
    @classmethod
    def validate_cost(cls, v):
        if v < 0:
            raise ValueError("cost must be non-negative")
        return v

    @field_validator("engagement_rate")
    @classmethod
    def validate_engagement(cls, v):
        if not (0 <= v <= 1):
            raise ValueError("engagement_rate must be between 0 and 1")
        return round(v, 4)

    @field_validator("fit_score")
    @classmethod
    def validate_fit(cls, v):
        if not (0 <= v <= 1):
            raise ValueError("fit_score must be between 0 and 1")
        return round(v, 4)


class KOLUpdate(BaseModel):
    """All fields optional for partial update."""
    name: Optional[str] = None
    country: Optional[str] = None
    category: Optional[str] = None
    followers: Optional[int] = None
    engagement_rate: Optional[float] = None
    fit_score: Optional[float] = None
    commission_rate: Optional[float] = None
    cost: Optional[float] = None
    avg_views: Optional[int] = None
    avg_likes: Optional[int] = None
    gender_ratio: Optional[float] = None
    age_group: Optional[str] = None
    tiktok_url: Optional[str] = None


# ════════════════════════════════════════════════════════════════
#  Helpers
# ════════════════════════════════════════════════════════════════
def _load() -> List[dict]:
    path = tenancy.data_path()
    if not os.path.exists(path):
        return []
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def _save(data: List[dict]):
    with open(tenancy.data_path(), "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def _next_id(data: List[dict]) -> int:
    if not data:
        return 1
    return max(d["id"] for d in data) + 1


# ════════════════════════════════════════════════════════════════
#  Flexible import helpers (CSV + Excel, FastMoss-friendly)
# ════════════════════════════════════════════════════════════════
def _norm_header(h) -> str:
    """Normalise a column header to snake_case ascii for alias matching."""
    return re.sub(r"[^a-z0-9]+", "_", str(h).strip().lower()).strip("_")


# Map many real-world / FastMoss header variants onto our canonical fields.
_FIELD_ALIASES = {
    "name":            {"name", "creator", "creator_name", "nickname", "username", "user_name", "handle", "account", "kol", "kol_name", "influencer", "display_name", "creator_nickname"},
    "country":         {"country", "region", "market", "location", "geo", "nation"},
    "category":        {"category", "niche", "industry", "vertical", "type", "content_category", "main_category", "tag", "tags"},
    "followers":       {"followers", "follower", "follower_count", "followers_count", "fans", "fans_count", "fans_number"},
    "engagement_rate": {"engagement_rate", "engagement", "er", "eng_rate", "avg_engagement_rate", "engagement_rate_percent", "interaction_rate"},
    "cost":            {"cost", "price", "quote", "fee", "rate_card", "collaboration_price", "collab_price", "unit_price", "price_usd", "cooperation_price", "ad_price"},
    "fit_score":       {"fit_score", "fit", "relevance", "relevance_score", "match_score", "brand_fit"},
    "commission_rate": {"commission_rate", "commission", "commission_pct"},
    "avg_views":       {"avg_views", "average_views", "views", "avg_view", "median_views", "avg_video_views"},
    "avg_likes":       {"avg_likes", "average_likes", "likes", "avg_like", "median_likes"},
    "gender_ratio":    {"gender_ratio", "female_ratio", "female", "female_percent", "female_pct", "pct_female", "audience_female"},
    "age_group":       {"age_group", "age", "age_range", "main_age", "age_bracket"},
    "tiktok_url":      {"tiktok_url", "url", "link", "profile", "profile_url", "tiktok", "tiktok_link", "homepage"},
}
_ALIAS_TO_FIELD = {alias: field for field, aliases in _FIELD_ALIASES.items() for alias in aliases}

_COUNTRY_NAME_TO_CODE = {
    "malaysia": "MY", "my": "MY",
    "indonesia": "ID", "id": "ID",
    "thailand": "TH", "th": "TH",
    "philippines": "PH", "the_philippines": "PH", "ph": "PH",
    "singapore": "SG", "sg": "SG",
    "vietnam": "VN", "viet_nam": "VN", "vn": "VN",
}
_CATEGORY_SYNONYMS = {
    "beauty": "beauty", "cosmetics": "beauty", "makeup": "beauty", "skincare": "beauty",
    "personal_care": "beauty", "beauty_personal_care": "beauty",
    "fashion": "fashion", "apparel": "fashion", "clothing": "fashion", "lifestyle": "fashion", "accessories": "fashion",
    "home": "home", "home_living": "home", "home_garden": "home", "household": "home", "furniture": "home", "home_kitchen": "home",
    "fmcg": "fmcg", "food": "fmcg", "grocery": "fmcg", "beverage": "fmcg", "snacks": "fmcg", "food_beverage": "fmcg",
}


def _parse_number(v):
    """Parse '250,000', '$1,200', '250K', '1.2M' → float. Returns None if blank."""
    if v is None:
        return None
    s = str(v).strip().replace(",", "").replace("$", "").replace("USD", "").strip()
    if s == "" or s.lower() in ("nan", "none", "null", "-", "n/a"):
        return None
    m = re.match(r"^([0-9]*\.?[0-9]+)\s*([kKmMbB]?)$", s)
    if m:
        mult = {"k": 1e3, "m": 1e6, "b": 1e9}.get(m.group(2).lower(), 1.0)
        return float(m.group(1)) * mult
    s2 = re.sub(r"[^0-9.\-]", "", s)
    try:
        return float(s2) if s2 not in ("", "-", ".") else None
    except ValueError:
        return None


def _parse_rate(v):
    """Parse a 0-1 rate from '8%', '8', or '0.08'. Values > 1 are treated as percentages."""
    if v is None or str(v).strip() == "":
        return None
    raw = str(v)
    n = _parse_number(raw.replace("%", ""))
    if n is None:
        return None
    if "%" in raw or n > 1.0:
        n = n / 100.0
    return n


def _estimate_cost(followers: int) -> float:
    """Rough collaboration-fee estimate when the source has no price (e.g. FastMoss).

    Tier-based per-follower rate card; flagged to the user so they can replace
    it with real negotiated quotes.
    """
    f = followers or 0
    rate = 0.003 if f >= 1_000_000 else 0.006 if f >= 100_000 else 0.012 if f >= 10_000 else 0.02
    return round(max(50.0, f * rate), 2)


def _read_csv_rows(content: bytes):
    try:
        text = content.decode("utf-8-sig")          # handle Excel BOM
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    reader = csv.DictReader(io.StringIO(text))
    return list(reader.fieldnames or []), [dict(r) for r in reader]


def _read_excel_rows(content: bytes):
    """Read the first worksheet of an .xlsx into (headers, list-of-dicts)."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    headers, rows = None, []
    for r in ws.iter_rows(values_only=True):
        is_blank = r is None or all(c is None or str(c).strip() == "" for c in r)
        if headers is None:
            if is_blank:
                continue
            headers = ["" if c is None else str(c) for c in r]
            continue
        if is_blank:
            continue
        rows.append({headers[i]: r[i] for i in range(min(len(headers), len(r)))})
    wb.close()
    return headers or [], rows


def _canonicalize_rows(headers, rows):
    """Remap each raw header to our canonical field name via the alias table."""
    mapping = {h: _ALIAS_TO_FIELD.get(_norm_header(h), _norm_header(h)) for h in headers}
    out = []
    for r in rows:
        out.append({mapping.get(k, _norm_header(k)): v for k, v in r.items()})
    return out


def _ingest_rows(rows, default_country=None, default_category=None):
    """Validate + insert canonicalised rows. Returns (added, estimated_cost, errors, total)."""
    data = _load()
    dc = (default_country or "").strip().upper() or None
    dcat = (default_category or "").strip().lower() or None
    added, estimated, errors = 0, 0, []

    for i, row in enumerate(rows, start=2):
        try:
            name = str(row.get("name") or "").strip()
            if not name:
                errors.append(f"Row {i}: missing creator name")
                continue

            craw = str(row.get("country") or "").strip()
            country = _COUNTRY_NAME_TO_CODE.get(_norm_header(craw), craw.upper()) if craw else (dc or "")
            if country not in VALID_COUNTRIES:
                errors.append(f"Row {i} ({name}): unknown country '{craw or '—'}' (need one of {', '.join(VALID_COUNTRIES)})")
                continue

            catraw = str(row.get("category") or "").strip()
            category = _CATEGORY_SYNONYMS.get(_norm_header(catraw), catraw.lower()) if catraw else (dcat or "")
            if category not in VALID_CATEGORIES:
                errors.append(f"Row {i} ({name}): unknown category '{catraw or '—'}' (need one of {', '.join(VALID_CATEGORIES)})")
                continue

            followers = _parse_number(row.get("followers"))
            if followers is None:
                errors.append(f"Row {i} ({name}): missing followers")
                continue
            followers = int(followers)

            eng = _parse_rate(row.get("engagement_rate"))
            if eng is None:
                errors.append(f"Row {i} ({name}): missing engagement_rate")
                continue

            cost = _parse_number(row.get("cost"))
            cost_estimated = False
            if cost is None or cost <= 0:
                cost = _estimate_cost(followers)
                estimated += 1
                cost_estimated = True

            avg_views = _parse_number(row.get("avg_views"))
            avg_likes = _parse_number(row.get("avg_likes"))
            gender = _parse_rate(row.get("gender_ratio"))

            record = {
                "id":              _next_id(data),
                "name":            name,
                "country":         country,
                "category":        category,
                "followers":       followers,
                "engagement_rate": round(eng, 4),
                "fit_score":       round(_parse_rate(row.get("fit_score")) or 0.75, 4),
                "commission_rate": round(_parse_rate(row.get("commission_rate")) or 0.15, 4),
                "cost":            round(float(cost), 2),
                "avg_views":       int(avg_views) if avg_views else None,
                "avg_likes":       int(avg_likes) if avg_likes else None,
                "gender_ratio":    round(gender, 2) if gender is not None else None,
                "age_group":       str(row.get("age_group") or "").strip() or None,
                "tiktok_url":      str(row.get("tiktok_url") or "").strip() or None,
                "cost_estimated":  cost_estimated,
            }
            data.append(_auto_fill(record))
            added += 1
        except (ValueError, KeyError, TypeError) as e:
            errors.append(f"Row {i}: {e}")

    _save(data)
    return added, estimated, errors, len(data)


def _auto_fill(d: dict) -> dict:
    """Fill computed fields if not provided."""
    if d.get("avg_views") is None:
        d["avg_views"] = int(d["followers"] * 0.3)
    if d.get("avg_likes") is None:
        d["avg_likes"] = int(d["avg_views"] * d["engagement_rate"])
    if d.get("gender_ratio") is None:
        d["gender_ratio"] = 0.5
    if d.get("age_group") is None:
        d["age_group"] = "18-24"
    return d


# ════════════════════════════════════════════════════════════════
#  Endpoints
# ════════════════════════════════════════════════════════════════
@router.post("/kols/add")
def add_kol(kol: KOLCreate, _t: str = Depends(require_tenant)):
    """Add a single KOL to the database."""
    data = _load()
    new_id = _next_id(data)

    record = {"id": new_id, **kol.model_dump()}
    record = _auto_fill(record)
    data.append(record)
    _save(data)

    return {"message": f"KOL '{kol.name}' added with id {new_id}", "id": new_id, "kol": record}


@router.put("/kols/{kol_id}")
def update_kol(kol_id: int, updates: KOLUpdate, _t: str = Depends(require_tenant)):
    """Update an existing KOL. Snapshots current metrics before applying changes."""
    data = _load()
    idx = next((i for i, d in enumerate(data) if d["id"] == kol_id), None)
    if idx is None:
        raise HTTPException(status_code=404, detail=f"KOL with id {kol_id} not found")

    changes = updates.model_dump(exclude_none=True)
    if "country" in changes and changes["country"] not in VALID_COUNTRIES:
        raise HTTPException(status_code=422, detail=f"country must be one of: {', '.join(VALID_COUNTRIES)}")
    if "category" in changes:
        changes["category"] = changes["category"].lower()
        if changes["category"] not in VALID_CATEGORIES:
            raise HTTPException(status_code=422, detail=f"category must be one of: {', '.join(VALID_CATEGORIES)}")

    # A manually entered cost is a real quote → clear the "estimated" flag.
    if "cost" in changes and changes["cost"] and changes["cost"] > 0:
        changes["cost_estimated"] = False

    # Auto-snapshot BEFORE applying changes (track history)
    _record_snapshot(data[idx])

    data[idx].update(changes)
    _save(data)

    return {"message": f"KOL {kol_id} updated", "kol": data[idx]}


@router.get("/kols/{kol_id}/history")
def get_kol_history(kol_id: int, _t: str = Depends(require_tenant)):
    """Return all metric snapshots for a KOL, newest first."""
    data = _load()
    if not any(d["id"] == kol_id for d in data):
        raise HTTPException(status_code=404, detail=f"KOL with id {kol_id} not found")

    history = _load_history()
    entry = next((h for h in history if h["kol_id"] == kol_id), None)
    snapshots = list(reversed(entry["snapshots"])) if entry else []
    return {"kol_id": kol_id, "snapshots": snapshots, "count": len(snapshots)}


@router.post("/kols/{kol_id}/simulate-update")
def simulate_kol_update(kol_id: int, _t: str = Depends(require_tenant)):
    """
    Simulate a TikTok API metric refresh: apply realistic random drift to
    engagement_rate, fit_score, followers, and record a new snapshot.
    Useful for demos and testing trend visualisation without a real API.
    """
    data = _load()
    idx = next((i for i, d in enumerate(data) if d["id"] == kol_id), None)
    if idx is None:
        raise HTTPException(status_code=404, detail=f"KOL with id {kol_id} not found")

    kol = data[idx]
    rng = random.Random()   # fresh RNG each call → different values

    def drift(val: float, lo: float, hi: float, min_v: float, max_v: float) -> float:
        """Apply multiplicative drift within [min_v, max_v]."""
        factor = 1 + rng.uniform(lo, hi)
        return round(max(min_v, min(max_v, val * factor)), 4)

    old_eng = kol.get("engagement_rate", 0.08)
    old_fit = kol.get("fit_score", 0.75)
    old_fol = kol.get("followers", 100000)

    # Snapshot BEFORE applying drift (consistent with update_kol behavior)
    _record_snapshot(dict(kol))

    kol["engagement_rate"] = drift(old_eng, -0.08, +0.12, 0.01, 0.50)
    kol["fit_score"]        = drift(old_fit, -0.05, +0.08, 0.10, 1.00)
    kol["followers"]        = int(drift(old_fol, -0.02, +0.06, 1000, 50_000_000))
    kol["avg_views"]        = int(kol["followers"] * rng.uniform(0.20, 0.40))
    kol["avg_likes"]        = int(kol["avg_views"] * kol["engagement_rate"])

    _save(data)

    return {
        "message": "Simulated metric refresh applied",
        "kol_id": kol_id,
        "changes": {
            "engagement_rate": {"before": old_eng, "after": kol["engagement_rate"]},
            "fit_score":        {"before": old_fit, "after": kol["fit_score"]},
            "followers":        {"before": old_fol, "after": kol["followers"]},
        },
    }


@router.delete("/kols/{kol_id}")
def delete_kol(kol_id: int, _t: str = Depends(require_tenant)):
    """Delete a KOL from the database."""
    data = _load()
    before = len(data)
    data = [d for d in data if d["id"] != kol_id]
    if len(data) == before:
        raise HTTPException(status_code=404, detail=f"KOL with id {kol_id} not found")
    _save(data)
    return {"message": f"KOL {kol_id} deleted", "remaining": len(data)}


@router.post("/kols/import-csv")
@router.post("/kols/import")
async def import_file(
    file: UploadFile = File(...),
    default_country: Optional[str] = Form(None),
    default_category: Optional[str] = Form(None),
    _t: str = Depends(require_tenant),
):
    """
    Bulk import creators from a CSV or Excel (.xlsx) file.

    Column headers are matched flexibly (FastMoss / spreadsheet variants like
    "Follower Count", "Engagement Rate (%)", "Region" are understood). Numbers
    may be formatted ("250K", "1.2M", "$1,200", "8%").

    Required (per row, or supplied as a default): name, country, category,
    followers, engagement_rate. `cost` is estimated from followers when absent.

    Defaults: `default_country` / `default_category` fill rows that lack those
    columns. If not supplied, they are inferred from the file name
    (e.g. "Beauty_MY.xlsx" → category=beauty, country=MY).
    """
    filename = (file.filename or "").strip()
    lower = filename.lower()
    content = await file.read()

    if lower.endswith(".csv"):
        headers, raw_rows = _read_csv_rows(content)
    elif lower.endswith(".xlsx"):
        try:
            headers, raw_rows = _read_excel_rows(content)
        except Exception as e:
            raise HTTPException(status_code=422, detail=f"Could not read Excel file: {e}")
    elif lower.endswith(".xls"):
        raise HTTPException(status_code=422, detail="Legacy .xls is not supported — please re-save as .xlsx or .csv.")
    else:
        raise HTTPException(status_code=422, detail="File must be a .csv or .xlsx file")

    if not raw_rows:
        raise HTTPException(status_code=422, detail="No data rows found in the file.")

    rows = _canonicalize_rows(headers, rows=raw_rows)

    # Infer per-file defaults from the file name when not explicitly provided.
    stem_tokens = set(_norm_header(os.path.splitext(filename)[0]).split("_"))
    if not default_category:
        default_category = next((_CATEGORY_SYNONYMS[t] for t in stem_tokens if t in _CATEGORY_SYNONYMS), None)
    if not default_country:
        default_country = next((_COUNTRY_NAME_TO_CODE[t] for t in stem_tokens if t in _COUNTRY_NAME_TO_CODE), None)

    added, estimated, errors, total = _ingest_rows(rows, default_country, default_category)

    msg = f"Imported {added} creators"
    if estimated:
        msg += f" · {estimated} had no price (cost estimated from followers)"
    return {
        "message": msg,
        "added": added,
        "estimated_cost": estimated,
        "applied_defaults": {"country": default_country, "category": default_category},
        "errors": errors,
        "total": total,
    }


@router.post("/kols/backfill-costs")
async def backfill_costs(file: UploadFile = File(...), _t: str = Depends(require_tenant)):
    """
    Bulk-update real collaboration prices from a CSV/Excel file.

    Each row identifies an existing creator — by `id`, `tiktok_url`, or `name`
    (matched in that priority) — and supplies a real `cost`/`price`. Matched
    creators get the real price and are marked `cost_estimated = false`, clearing
    the "est." badge. Rows with no price or no match are reported back.
    """
    filename = (file.filename or "").lower()
    content = await file.read()
    if filename.endswith(".csv"):
        headers, raw_rows = _read_csv_rows(content)
    elif filename.endswith(".xlsx"):
        try:
            headers, raw_rows = _read_excel_rows(content)
        except Exception as e:
            raise HTTPException(status_code=422, detail=f"Could not read Excel file: {e}")
    else:
        raise HTTPException(status_code=422, detail="File must be a .csv or .xlsx file")

    if not raw_rows:
        raise HTTPException(status_code=422, detail="No data rows found in the file.")

    rows = _canonicalize_rows(headers, raw_rows)
    data = _load()
    by_id = {d["id"]: d for d in data}
    by_url, by_name = {}, {}
    for d in data:
        url = (d.get("tiktok_url") or "").strip().lower()
        if url:
            by_url.setdefault(url, d)
        by_name.setdefault((d.get("name") or "").strip().lower(), d)

    updated, no_price, unmatched = 0, 0, []
    for i, row in enumerate(rows, start=2):
        cost = _parse_number(row.get("cost"))
        if cost is None or cost <= 0:
            no_price += 1
            continue

        target = None
        rid = _parse_number(row.get("id")) or _parse_number(row.get("creator_id"))
        if rid is not None:
            target = by_id.get(int(rid))
        if target is None:
            target = by_url.get((row.get("tiktok_url") or "").strip().lower())
        if target is None:
            target = by_name.get((row.get("name") or "").strip().lower())
        if target is None:
            unmatched.append(str(row.get("name") or row.get("tiktok_url") or f"row {i}"))
            continue

        target["cost"] = round(float(cost), 2)
        target["cost_estimated"] = False
        updated += 1

    _save(data)
    msg = f"Updated {updated} creator prices"
    if unmatched:
        msg += f" · {len(unmatched)} rows didn't match any creator"
    return {
        "message": msg,
        "updated": updated,
        "no_price_rows": no_price,
        "unmatched": unmatched[:50],
        "unmatched_count": len(unmatched),
    }


@router.get("/kols/template")
def download_template():
    """Return a CSV template with headers and one example row."""
    from fastapi.responses import StreamingResponse

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(CSV_TEMPLATE_HEADER)
    writer.writerow([
        "Lisa Beauty", "MY", "beauty", 250000,
        0.08, 0.85, 0.15, 1200,
        75000, 6000, 0.78, "18-24",
        "https://tiktok.com/@lisa_beauty",
    ])
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=kol_template.csv"},
    )


@router.get("/kols/template-excel")
def download_template_excel():
    """Return an .xlsx template with headers and one example row."""
    import openpyxl
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "creators"
    ws.append(CSV_TEMPLATE_HEADER)
    ws.append([
        "Lisa Beauty", "MY", "beauty", 250000,
        0.08, 0.85, 0.15, 1200,
        75000, 6000, 0.78, "18-24",
        "https://tiktok.com/@lisa_beauty",
    ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=kol_template.xlsx"},
    )


@router.get("/kols/export")
def export_csv(_t: str = Depends(require_tenant)):
    """Export all KOLs as a downloadable CSV file."""
    from fastapi.responses import StreamingResponse

    data = _load()
    output = io.StringIO()
    if data:
        all_keys = []
        seen = set()
        for d in data:
            for k in d.keys():
                if k not in seen:
                    all_keys.append(k)
                    seen.add(k)
        writer = csv.DictWriter(output, fieldnames=all_keys, extrasaction="ignore")
        writer.writeheader()
        for row in data:
            writer.writerow({k: row.get(k, "") for k in all_keys})
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=kol_export.csv"},
    )


@router.post("/kols/reset")
def reset_database(_t: str = Depends(require_tenant)):
    """
    Restore the library to this tenant's seed baseline (seed_kols.json) —
    the initialization creator library every merchant starts from.

    A merchant can freely add / edit / delete / import while experimenting,
    then reset to snap back to the curated seed. If no seed baseline exists,
    this clears the library to empty instead.

    Also clears metric history so reset state is clean (snapshots that
    referenced edited/removed creators do not linger).
    """
    seed_p = tenancy.seed_path()
    if os.path.exists(seed_p):
        with open(seed_p, encoding="utf-8") as f:
            seed = json.load(f)
        _save(seed)
        _save_history([])
        return {"message": "Library restored to seed baseline", "total": len(seed)}

    _save([])
    _save_history([])
    return {"message": "Database cleared successfully", "total": 0}